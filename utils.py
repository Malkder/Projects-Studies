"""
Utility functions for the performance attribution reporting app.

This module centralizes helper functions to load the mapping file, parse
Bloomberg BONE exports, build the analysis tables, restructure columns to
create multi‑level headers, generate Excel downloads and integrate IFM
data from either an SQL server or an uploaded Excel file.

Key features supported:

• Mapping: reads a CSV containing at least `code_report`, `ID`, `NOM`,
  `DEVISE`, with optional `TYPE_PART`, `PORTEFEUILLE` and `BENCHMARK`.
  Adds a lowercase helper column `code_report_lower` for case‑insensitive
  matching.

• BONE parsing: extracts performance values from the second worksheet of
  a BONE export and returns PF_BONE and BM_BONE for a given period.
  The parser silently ignores files whose names do not contain
  `_<period>_` (e.g. `_MTD_` when the user selects MTD).

• Table construction: produces the "Analyse interne" and "Attribution"
  tables. These tables include the optional IFM columns (PF_IFM,
  BM_IFM) but leave them blank until filled later. They also compute
  preliminary BONE‑based differences.

• Table restructuring: converts flat DataFrames into ones with
  hierarchical column indices to group Portefeuille, Benchmark and
  Écarts columns under one parent header in the UI.

• Excel export: writes a DataFrame to an in‑memory Excel file without
  changing numeric formatting. Each column is given a comfortable
  width. The function returns the bytes and a suggested filename.

• IFM integration: defines helper functions to merge IFM data into the
  internal and attribution tables. A function `get_ifm_values` is
  provided as an example for SQL access via `pyodbc`. Additional
  helpers exist to merge a list of (ID, PF_IFM, BM_IFM) tuples into
  your tables and compute the relevant differences.

Note that this module is deliberately agnostic about the UI. All
parameters required for mapping, file parsing, period filtering and
querying IFM are passed in by the Streamlit application. If you need
to customize column names or period mappings for IFM, adjust the
PERIOD_MAP and SQL in `get_ifm_values` accordingly.
"""

from __future__ import annotations

import io
import os
from datetime import date
from typing import Dict, List, Optional, Tuple, Any

import pandas as pd
from openpyxl import load_workbook
import numpy as np


# Map the UI period to the IFM period value.
PERIOD_MAP: Dict[str, str] = {
    "MTD": "1MOIS",
    "QTD": "3MOIS",
    "YTD": "DEB_ANNEE",
}

# -----------------------------
# Helpers
# -----------------------------

def _to_float(val: Any) -> Optional[float]:
    """Convert flexible numeric strings to float.
    Handles values like "1,23%", "-0.45%", "(1,2%)", "1.23", 0.0123.
    Returns None if conversion fails.
    """
    if val is None:
        return None
    # Already numeric
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip()
    if not s:
        return None
    # Parentheses negative
    neg = s.startswith("(") and s.endswith(")")
    if neg:
        s = s[1:-1]
    # Remove percent sign
    pct = s.endswith("%")
    if pct:
        s = s[:-1]
    # Replace comma decimal with dot
    s = s.replace(" ", "").replace("\xa0", "").replace(",", ".")
    try:
        num = float(s)
        if pct:
            num = num / 100.0
        if neg:
            num = -num
        return num
    except Exception:
        return None


def _flatten_columns_for_excel(df: pd.DataFrame) -> pd.DataFrame:
    """If df has MultiIndex columns, flatten them for Excel export."""
    if isinstance(df.columns, pd.MultiIndex):
        flat = [
            (str(a) if a else "") + (" – " + str(b) if b else "")
            for a, b in df.columns
        ]
        new_df = df.copy()
        new_df.columns = flat
        return new_df
    return df


def load_mapping(data: Optional[bytes]) -> pd.DataFrame:
    """
    Load the mapping CSV from a bytes object. The function accepts
    variations in column names by stripping whitespace and matching
    insensitive keys. Expected columns include at least:

    - code_report
    - ID
    - NOM
    - DEVISE

    Optional columns that improve functionality are:

    - TYPE_PART (used for IFM matching)
    - PORTEFEUILLE
    - BENCHMARK

    A helper column `code_report_lower` is added to enable
    case‑insensitive matching with BONE exports.
    """
    if data is None:
        return pd.DataFrame()
    # Try CSV with auto sep, then fallback to semicolon
    try:
        df = pd.read_csv(io.BytesIO(data))
    except Exception:
        try:
            df = pd.read_csv(io.BytesIO(data), sep=';')
        except Exception:
            return pd.DataFrame()

    # Normalize header names by stripping whitespace
    df.columns = [str(c).strip() for c in df.columns]

    # Define possible mappings to canonical names
    rename_map: Dict[str, str] = {
        "Code_Report": "code_report", "CODE_REPORT": "code_report", "code_report": "code_report",
        "Id": "ID", "id": "ID", "ID": "ID",
        "Nom": "NOM", "nom": "NOM", "NOM": "NOM",
        "Devise": "DEVISE", "devise": "DEVISE", "DEVISE": "DEVISE",
        "Type Part": "TYPE_PART", "TYPE PART": "TYPE_PART", "TYPE_PART": "TYPE_PART", "type_part": "TYPE_PART",
        "Portefeuille": "PORTEFEUILLE", "portefeuille": "PORTEFEUILLE", "PORTEFEUILLE": "PORTEFEUILLE",
        "Benchmark": "BENCHMARK", "benchmark": "BENCHMARK", "BENCHMARK": "BENCHMARK",
    }
    canonical_cols: Dict[str, str] = {}
    for col in df.columns:
        if col in rename_map:
            canonical_cols[col] = rename_map[col]
        else:
            canonical_cols[col] = col  # leave unknown columns unchanged
    df = df.rename(columns=canonical_cols)

    # Strip spaces in key text columns to stabilise matching
    for _c in ["code_report", "ID", "NOM", "DEVISE", "TYPE_PART", "PORTEFEUILLE", "BENCHMARK"]:
        if _c in df.columns:
            df[_c] = df[_c].astype(str).str.strip()

    # Ensure required columns exist
    required = {"code_report", "ID", "NOM", "DEVISE"}
    if not required.issubset(set(df.columns)):
        return pd.DataFrame()

    # Add missing optional columns
    for col in ["TYPE_PART", "PORTEFEUILLE", "BENCHMARK"]:
        if col not in df.columns:
            df[col] = ""

    # Ensure ID is string to avoid merge type mismatches
    if "ID" in df.columns:
        df["ID"] = df["ID"].astype(str)

    # Lowercase helper for matching
    df["code_report_lower"] = df["code_report"].astype(str).str.lower()
    return df


def parse_bone_file(file_obj: object, period: str) -> Tuple[Optional[str], Optional[float], Optional[float], List[str]]:
    """
    Parse a single Bloomberg BONE export.

    Expected file name pattern: `<code_report>_<PERIOD>_<MMYYYY>.xlsx`.
    We only process files whose names contain `_{PERIOD}_` (e.g. `_MTD_`).

    Values location (fixed): second worksheet (index 1),
    - PTF_BONE  in cell C9
    - BENCH_ONE in cell C10

    Returns
    -------
    (code_report, pf_bone, bm_bone, warnings)
    where pf_bone is the portfolio value (C9) and bm_bone is the benchmark (C10).
    """
    warnings: List[str] = []
    code_report: Optional[str] = None

    # 1) Filter by period using the file name and extract code_report
    try:
        filename = os.path.basename(file_obj.name)  # type: ignore[attr-defined]
        if f"_{period}_" not in filename:
            return None, None, None, []
        code_report = filename.split(f"_{period}_")[0]
    except Exception:
        # If no filename is available, we cannot verify the period reliably
        warnings.append("Nom de fichier indisponible; impossible de vérifier la période.")

    # 2) Read workbook bytes
    try:
        data_bytes = file_obj.read()
    except Exception as e:
        warnings.append(f"Erreur lecture du fichier: {e}")
        return code_report, None, None, warnings

    try:
        wb = load_workbook(filename=io.BytesIO(data_bytes), data_only=True)
        if len(wb.sheetnames) < 2:
            warnings.append("Le fichier ne contient pas de 2ᵉ onglet.")
            return code_report, None, None, warnings
        ws = wb[wb.sheetnames[1]]  # second worksheet

        # 3) Fixed cells: C9 (PTF_BONE), C10 (BENCH_ONE)
        pf_raw = ws["C9"].value if ws["C9"] is not None else None
        bm_raw = ws["C10"].value if ws["C10"] is not None else None

        pf = _to_float(pf_raw)
        bm = _to_float(bm_raw)
        if pf is None:
            warnings.append("Cellule C9 (PTF_BONE) introuvable ou non numérique.")
        if bm is None:
            warnings.append("Cellule C10 (BENCH_ONE) introuvable ou non numérique.")
        return code_report, pf, bm, warnings
    except Exception as e:
        warnings.append(f"Erreur ouverture xlsx: {e}")
        return code_report, None, None, warnings


def build_internal_table(
    mapping_df: pd.DataFrame, results: List[Tuple[str, Optional[float], Optional[float], List[str]]]
) -> pd.DataFrame:
    """
    Build the internal performance analysis table from mapping and BONE
    results. Each row corresponds to one `code_report`, matched on
    `code_report_lower` against the mapping. If no mapping is found,
    fallback metadata is filled with `N/A` or blanks. The columns
    PF_IFM, BM_IFM and their differences remain None until injected by
    an external process.
    """
    rows: List[Dict[str, Any]] = []
    for code_report, pf_bone, bm_bone, _ in results:
        if code_report is None:
            continue
        match = mapping_df[mapping_df["code_report_lower"] == code_report.lower()]
        pf_b = pf_bone if pf_bone is not None else np.nan
        bm_b = bm_bone if bm_bone is not None else np.nan
        if match.empty:
            rows.append({
                "ID": "N/A", "NOM": code_report, "DEVISE": "", "TYPE_PART": "",
                "PORTEFEUILLE": "", "BENCHMARK": "",
                "PF_IFM": np.nan, "PF_BONE": pf_b, "BM_IFM": np.nan, "BM_BONE": bm_b,
                "PF_IFM–PF_BONE": np.nan, "BM_IFM–BM_BONE": np.nan,
            })
        else:
            r = match.iloc[0]
            rows.append({
                "ID": r["ID"], "NOM": r["NOM"], "DEVISE": r["DEVISE"], "TYPE_PART": r["TYPE_PART"],
                "PORTEFEUILLE": r["PORTEFEUILLE"], "BENCHMARK": r["BENCHMARK"],
                "PF_IFM": np.nan, "PF_BONE": pf_b, "BM_IFM": np.nan, "BM_BONE": bm_b,
                "PF_IFM–PF_BONE": np.nan, "BM_IFM–BM_BONE": np.nan,
            })
    cols = [
        "ID", "NOM", "DEVISE", "TYPE_PART", "PORTEFEUILLE", "BENCHMARK",
        "PF_IFM", "PF_BONE", "BM_IFM", "BM_BONE", "PF_IFM–PF_BONE", "BM_IFM–BM_BONE",
    ]
    df_out = pd.DataFrame(rows, columns=cols)
    for c in ["PF_IFM", "PF_BONE", "BM_IFM", "BM_BONE", "PF_IFM–PF_BONE", "BM_IFM–BM_BONE"]:
        if c in df_out.columns:
            df_out[c] = pd.to_numeric(df_out[c], errors="coerce")
    return df_out


def build_attribution_table(
    mapping_df: pd.DataFrame, results: List[Tuple[str, Optional[float], Optional[float], List[str]]]
) -> pd.DataFrame:
    """
    Build the attribution table (portefeuille vs benchmark) from
    mapping and BONE results. Each row corresponds to a matched
    `code_report`. The BONE difference (PF_BONE – BM_BONE) is
    precomputed. The IFM columns remain None until injection.
    """
    rows: List[Dict[str, Any]] = []
    for code_report, pf_bone, bm_bone, _ in results:
        if code_report is None:
            continue
        diff = (pf_bone - bm_bone) if (pf_bone is not None and bm_bone is not None) else np.nan
        match = mapping_df[mapping_df["code_report_lower"] == code_report.lower()]
        pf_b = pf_bone if pf_bone is not None else np.nan
        bm_b = bm_bone if bm_bone is not None else np.nan
        if match.empty:
            rows.append({
                "ID": "N/A", "NOM": code_report, "DEVISE": "", "TYPE_PART": "",
                "PORTEFEUILLE": "", "BENCHMARK": "",
                "PF_IFM": np.nan, "PF_BONE": pf_b, "BM_IFM": np.nan, "BM_BONE": bm_b,
                "PF_IFM–BM_IFM": np.nan, "PF_BONE–BM_BONE": diff,
            })
        else:
            r = match.iloc[0]
            rows.append({
                "ID": r["ID"], "NOM": r["NOM"], "DEVISE": r["DEVISE"], "TYPE_PART": r["TYPE_PART"],
                "PORTEFEUILLE": r["PORTEFEUILLE"], "BENCHMARK": r["BENCHMARK"],
                "PF_IFM": np.nan, "PF_BONE": pf_b, "BM_IFM": np.nan, "BM_BONE": bm_b,
                "PF_IFM–BM_IFM": np.nan, "PF_BONE–BM_BONE": diff,
            })
    cols = [
        "ID", "NOM", "DEVISE", "TYPE_PART", "PORTEFEUILLE", "BENCHMARK",
        "PF_IFM", "PF_BONE", "BM_IFM", "BM_BONE", "PF_IFM–BM_IFM", "PF_BONE–BM_BONE",
    ]
    df_out = pd.DataFrame(rows, columns=cols)
    for c in ["PF_IFM", "PF_BONE", "BM_IFM", "BM_BONE", "PF_IFM–BM_IFM", "PF_BONE–BM_BONE"]:
        if c in df_out.columns:
            df_out[c] = pd.to_numeric(df_out[c], errors="coerce")
    return df_out


def restructure_internal_df(df: pd.DataFrame) -> pd.DataFrame:
    """
    Convert the internal DataFrame into one with a MultiIndex column
    hierarchy suitable for display. The Portefeuille columns (PF_IFM,
    PF_BONE), Benchmark columns (BM_IFM, BM_BONE) and Écarts columns
    (PF_IFM–PF_BONE, BM_IFM–BM_BONE) are grouped under their respective
    parent headers. All other metadata columns remain top‑level.
    """
    col_tuples: List[Tuple[str, str]] = []
    for col in df.columns:
        if col in {"PF_IFM", "PF_BONE"}:
            col_tuples.append(("PORTFEUILLE", col.split("_")[-1]))
        elif col in {"BM_IFM", "BM_BONE"}:
            col_tuples.append(("BENCHMARK", col.split("_")[-1]))
        elif col in {"PF_IFM–PF_BONE", "BM_IFM–BM_BONE"}:
            col_tuples.append(("ECARTS", col))
        else:
            col_tuples.append((col, ""))
    new_df = df.copy()
    new_df.columns = pd.MultiIndex.from_tuples(col_tuples)
    return new_df


def restructure_attribution_df(df: pd.DataFrame) -> pd.DataFrame:
    """
    Convert the attribution DataFrame into a MultiIndex form similar
    to the internal table. The Portefeuille (PF_IFM, PF_BONE),
    Benchmark (BM_IFM, BM_BONE) and Écarts (PF_IFM–BM_IFM,
    PF_BONE–BM_BONE) columns are grouped. Other columns remain
    top‑level.
    """
    col_tuples: List[Tuple[str, str]] = []
    for col in df.columns:
        if col in {"PF_IFM", "PF_BONE"}:
            col_tuples.append(("PORTFEUILLE", col.split("_")[-1]))
        elif col in {"BM_IFM", "BM_BONE"}:
            col_tuples.append(("BENCHMARK", col.split("_")[-1]))
        elif col in {"PF_IFM–BM_IFM", "PF_BONE–BM_BONE"}:
            col_tuples.append(("ECARTS", col))
        else:
            col_tuples.append((col, ""))
    new_df = df.copy()
    new_df.columns = pd.MultiIndex.from_tuples(col_tuples)
    return new_df


def generate_excel_download(df: pd.DataFrame, filename: str) -> Tuple[bytes, str]:
    """
    Write a DataFrame to an Excel file in memory. The function does not
    apply any numeric formatting; decimals remain decimals. All
    columns receive a fixed width for better readability. Returns the
    bytes and the suggested filename.
    """
    output = io.BytesIO()
    # Flatten columns if MultiIndex to avoid awkward headers in Excel
    df_to_write = _flatten_columns_for_excel(df)
    with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
        df_to_write.to_excel(writer, index=False, sheet_name="Data")
        worksheet = writer.sheets["Data"]
        for i, _ in enumerate(df_to_write.columns):
            worksheet.set_column(i, i, 18)
        # writer.close()  # Removed: Context manager already closes the writer
    output.seek(0)
    return output.read(), filename


def get_ifm_values(
    conn_str: str,
    id_ptf: str,
    type_part: str,
    devise: str,
    period_ui: str,
    asof: date,
    date_column: str = "DT_PERF",
    table: str = "infoma4",
) -> Tuple[Optional[float], Optional[float]]:
    """
    Query the SQL database for IFM values for a given portfolio. This
    function uses `pyodbc` to connect. It selects columns
    `CS_PERF_BRUTE` (portefeuille) and `CS_PERF_BENCH` (benchmark) filtered by:

    - ID_PTF_CDCAM = id_ptf
    - TY_PORTEUR = type_part
    - PERIOD_CALCUL = PERIOD_MAP[period_ui]
    - ID_DEVISE_PERF = devise
    - DT_PERF = asof

    You can adjust `table` and `date_column` to suit your schema. If
    multiple rows match, only the first row is returned. Errors are
    silently ignored and None values returned.
    """
    PERIOD_CALCUL = PERIOD_MAP.get(period_ui, "1MOIS")
    query = f"""
    SELECT CS_PERF_BRUTE, CS_PERF_BENCH
    FROM {table}
    WHERE ID_PTF_CDCAM = ?
      AND TY_PORTEUR   = ?
      AND PERIOD_CALCUL = ?
      AND ID_DEVISE_PERF = ?
      AND {date_column} = ?
    """
    params = (id_ptf, type_part, PERIOD_CALCUL, devise, asof)
    pf, bm = None, None
    try:
        import pyodbc  # defer import until call
        with pyodbc.connect(conn_str) as cnx:
            cur = cnx.cursor()
            cur.execute(query, params)
            row = cur.fetchone()
            if row:
                pf = float(row[0]) if row[0] is not None else None
                bm = float(row[1]) if row[1] is not None else None
    except Exception as e:
        # Logging left to the caller
        print("IFM SQL error:", e)
    return pf, bm

# -----------------------------------------------------------------------------
# Bulk retrieval for IFM data (INFOMAD) via SQL
# -----------------------------------------------------------------------------


# -----------------------------------------------------------------------------
# Bulk retrieval with full criteria alignment (ID, TYPE, DEVISE, PERIOD)
# -----------------------------------------------------------------------------

def build_ifm_criteria(mapping_df: pd.DataFrame, period_ui: str) -> List[Tuple[str, str, str, str]]:
    """
    Build the list of IFM criteria tuples (ID_PTF_CDCAM, TY_PORTEUR, ID_DEVISE_PERF, PERIOD_CALCUL)
    from the mapping table and the UI period.
    Only rows with non-empty ID/TYPE/DEVISE are included.
    """
    period_calc = PERIOD_MAP.get(period_ui, "1MOIS")
    crit: List[Tuple[str, str, str, str]] = []
    for _, r in mapping_df.iterrows():
        idv = str(r.get("ID", "")).strip()
        typ = str(r.get("TYPE_PART", "")).strip()
        dev = str(r.get("DEVISE", "")).strip()
        if idv and typ and dev:
            crit.append((idv, typ, dev, period_calc))
    return crit


def fetch_ifm_data_bulk_criteria(
    conn_str: str,
    asof: date,
    criteria: List[Tuple[str, str, str, str]],
    table: str = "INFOMAD.dbo.IFM_PERF_PTF",
) -> pd.DataFrame:
    """
    Fetch IFM rows using FULL criteria alignment (same logic as get_ifm_values):
    - ID_PTF_CDCAM, TY_PORTEUR, ID_DEVISE_PERF, PERIOD_CALCUL, and DT_PERF.

    Parameters
    ----------
    conn_str : str
        ODBC connection string
    asof : date
        DT_PERF filter
    criteria : list of tuples
        Each tuple is (ID_PTF_CDCAM, TY_PORTEUR, ID_DEVISE_PERF, PERIOD_CALCUL)
    table : str
        Table name (default: INFOMAD.dbo.IFM_PERF_PTF)

    Returns
    -------
    DataFrame with at least the columns: ID_PTF_CDCAM, TY_PORTEUR, PERIOD_CALCUL,
    ID_DEVISE_PERF, CS_PERF_BRUTE, CS_PERF_BENCH.
    """
    import pyodbc
    if not criteria:
        return pd.DataFrame()

    # Build VALUES table for the criteria (parameterized)
    # SQL Server supports VALUES lists: (?,?,?,?),(...)
    values_rows = ",".join(["(?, ?, ?, ?)"] * len(criteria))
    query = (
        f"WITH crit(ID_PTF_CDCAM, TY_PORTEUR, ID_DEVISE_PERF, PERIOD_CALCUL) AS (VALUES {values_rows}) "
        f"SELECT i.ID_PTF_CDCAM, i.TY_PORTEUR, i.PERIOD_CALCUL, i.ID_DEVISE_PERF, "
        f"       i.CS_PERF_BRUTE, i.CS_PERF_BENCH "
        f"FROM {table} i "
        f"JOIN crit c ON i.ID_PTF_CDCAM = c.ID_PTF_CDCAM "
        f"           AND i.TY_PORTEUR   = c.TY_PORTEUR "
        f"           AND i.ID_DEVISE_PERF = c.ID_DEVISE_PERF "
        f"           AND i.PERIOD_CALCUL  = c.PERIOD_CALCUL "
        f"WHERE i.DT_PERF = ?"
    )

    # Flatten parameters: for each (id, type, dev, period) then the asof at the end
    flat_params: List[Any] = []
    for (idv, typ, dev, per) in criteria:
        flat_params.extend([idv, typ, dev, per])
    flat_params.append(asof)

    try:
        with pyodbc.connect(conn_str) as cnx:
            cur = cnx.cursor()
            cur.execute(query, flat_params)
            rows = cur.fetchall()
            cols = [c[0] for c in cur.description] if cur.description else []
    except Exception as e:
        print("IFM bulk (criteria) query error:", e)
        return pd.DataFrame()

    df = pd.DataFrame.from_records(rows, columns=cols)
    if not df.empty and "ID_PTF_CDCAM" in df.columns:
        df["ID_PTF_CDCAM"] = df["ID_PTF_CDCAM"].astype(str)
    return df


def pack_ifm_rows_from_df(df: pd.DataFrame) -> List[Tuple[str, Optional[float], Optional[float]]]:
    """
    Convert an IFM DataFrame (with CS_PERF_BRUTE / CS_PERF_BENCH) to the list of
    tuples expected by the merge functions: (ID, PF_IFM, BM_IFM).
    If duplicates per ID exist, keeps the first non-null values after sorting
    by nothing (stable fetch order); callers can pre-sort if needed.
    """
    if df.empty:
        return []
    keep_cols = [
        c for c in ["ID_PTF_CDCAM", "CS_PERF_BRUTE", "CS_PERF_BENCH"] if c in df.columns
    ]
    slim = df[keep_cols].copy()
    # Group by ID and take first non-null values
    agg = (
        slim.groupby("ID_PTF_CDCAM", as_index=False)
            .agg({"CS_PERF_BRUTE": "first", "CS_PERF_BENCH": "first"})
    )
    # Build tuples in expected order (ID, PF_IFM, BM_IFM)
    out: List[Tuple[str, Optional[float], Optional[float]]] = []
    for _idx, row in agg.iterrows():
        out.append((str(row["ID_PTF_CDCAM"]), row.get("CS_PERF_BRUTE"), row.get("CS_PERF_BENCH")))
    return out


def fetch_ifm_data_bulk(
    conn_str: str,
    asof: date,
    ids: List[str],
    table: str = "INFOMAD.dbo.IFM_PERF_PTF",
) -> pd.DataFrame:
    """
    Fetch IFM performance data for multiple portfolios at once from the
    specified table. This function is designed to query the
    `INFOMAD.dbo.IFM_PERF_PTF` table (or a table with the same
    structure) and return a DataFrame containing the key fields used for
    performance reporting.

    Parameters
    ----------
    conn_str : str
        Full ODBC connection string for pyodbc.
    asof : datetime.date
        Date of performance to filter (DT_PERF = asof).
    ids : List[str]
        List of portfolio IDs (ID_PTF_CDCAM) to retrieve. The query uses
        an `IN` clause with these values.
    table : str, default "INFOMAD.dbo.IFM_PERF_PTF"
        Name of the table to query. The default corresponds to the
        Infomad performance table. Modify if necessary.

    Returns
    -------
    pandas.DataFrame
        DataFrame containing at least the columns:
          - ID_PTF_CDCAM
          - TY_PORTEUR
          - PERIOD_CALCUL
          - ID_DEVISE_PERF
          - CS_PERF_BRUTE (portfolio performance)
          - CS_PERF_BENCH (benchmark performance)
    If the query fails or returns no rows, an empty DataFrame is
    returned.

    Notes
    -----
    This function constructs a parameterised SQL query to avoid SQL
    injection and to handle dynamic lists of IDs. It uses `pyodbc`
    internally and catches exceptions gracefully.
    """
    import pyodbc
    # Build placeholder string for IN clause
    if not ids:
        return pd.DataFrame()
    placeholders = ",".join(["?"] * len(ids))
    query = (
        f"SELECT ID_PTF_CDCAM, TY_PORTEUR, PERIOD_CALCUL, ID_DEVISE_PERF, "
        f"CS_PERF_BRUTE, CS_PERF_BENCH "
        f"FROM {table} "
        f"WHERE DT_PERF = ? AND ID_PTF_CDCAM IN ({placeholders})"
    )
    params = [asof] + ids
    try:
        with pyodbc.connect(conn_str) as cnx:
            cur = cnx.cursor()
            cur.execute(query, params)
            rows = cur.fetchall()
            cols = [c[0] for c in cur.description] if cur.description else []
    except Exception as e:
        print("IFM bulk query error:", e)
        return pd.DataFrame()
    df_out = pd.DataFrame.from_records(rows, columns=cols)
    if not df_out.empty and "ID_PTF_CDCAM" in df_out.columns:
        df_out["ID_PTF_CDCAM"] = df_out["ID_PTF_CDCAM"].astype(str)
    return df_out


def merge_ifm_on_internal(
    internal_df: pd.DataFrame, ifm_list: List[Tuple[str, Optional[float], Optional[float]]]
) -> pd.DataFrame:
    """
    Merge a list of (ID, PF_IFM, BM_IFM) tuples into the internal
    analysis DataFrame. The merge is performed on the ID column,
    preserving rows. After merging, the differences PF_IFM–PF_BONE and
    BM_IFM–BM_BONE are recalculated.
    """
    ifm_df = pd.DataFrame(ifm_list, columns=["ID", "PF_IFM", "BM_IFM"])
    internal_df = internal_df.copy()
    if "ID" in internal_df.columns:
        internal_df["ID"] = internal_df["ID"].astype(str)
    if "ID" in ifm_df.columns:
        ifm_df["ID"] = ifm_df["ID"].astype(str)
    merged = internal_df.drop(columns=["PF_IFM", "BM_IFM"], errors="ignore").merge(
        ifm_df, on="ID", how="left"
    )
    merged["PF_IFM–PF_BONE"] = merged.apply(
        lambda r: (r["PF_IFM"] - r["PF_BONE"]) if pd.notnull(r["PF_IFM"]) and pd.notnull(r["PF_BONE"]) else None,
        axis=1,
    )
    merged["BM_IFM–BM_BONE"] = merged.apply(
        lambda r: (r["BM_IFM"] - r["BM_BONE"]) if pd.notnull(r["BM_IFM"]) and pd.notnull(r["BM_BONE"]) else None,
        axis=1,
    )
    # Ensure numeric dtypes for tables/formatters (avoid warnings about non-numeric)
    for c in ["PF_IFM", "BM_IFM", "PF_BONE", "BM_BONE", "PF_IFM–PF_BONE", "BM_IFM–BM_BONE"]:
        if c in merged.columns:
            merged[c] = pd.to_numeric(merged[c], errors="coerce")
    return merged


def merge_ifm_on_attribution(
    attribution_df: pd.DataFrame, ifm_list: List[Tuple[str, Optional[float], Optional[float]]]
) -> pd.DataFrame:
    """
    Merge a list of (ID, PF_IFM, BM_IFM) tuples into the attribution
    table. After merging, compute PF_IFM–BM_IFM. The BONE difference
    remains unchanged.
    """
    ifm_df = pd.DataFrame(ifm_list, columns=["ID", "PF_IFM", "BM_IFM"])
    # Harmonize key type to string on both sides (defensive for Excel/SQL mixed inputs)
    attribution_df = attribution_df.copy()
    if "ID" in attribution_df.columns:
        attribution_df["ID"] = attribution_df["ID"].astype(str)
    if "ID" in ifm_df.columns:
        ifm_df["ID"] = ifm_df["ID"].astype(str)

    # Drop placeholders to avoid overlap, then merge
    merged = attribution_df.drop(columns=["PF_IFM", "BM_IFM"], errors="ignore").merge(
        ifm_df, on="ID", how="left"
    )
    merged["PF_IFM–BM_IFM"] = merged.apply(
        lambda r: (r["PF_IFM"] - r["BM_IFM"]) if pd.notnull(r["PF_IFM"]) and pd.notnull(r["BM_IFM"]) else None,
        axis=1,
    )
    # Ensure numeric dtypes for tables/formatters (avoid warnings about non-numeric)
    for c in ["PF_IFM", "BM_IFM", "PF_BONE", "BM_BONE", "PF_IFM–BM_IFM", "PF_BONE–BM_BONE"]:
        if c in merged.columns:
            merged[c] = pd.to_numeric(merged[c], errors="coerce")
    return merged

# -----------------------------------------------------------------------------
# NOTES (test VS Code refresh)
# -----------------------------------------------------------------------------
# Ce bloc est ajouté pour vérifier que les changements sont bien pris en compte
# par VS Code / Streamlit. Il n’impacte pas l’exécution du module.
NOTES_TEST = "Probe: modifications appliquées dans utils.py (test VS Code)."

def __utils_change_probe__() -> str:
    """Retourne une chaîne indicative confirmant la mise à jour du fichier."""
    return NOTES_TEST

# -----------------------------------------------------------------------------
# Display helpers: produce fixed, user-facing column sets
# -----------------------------------------------------------------------------

def to_internal_display(df: pd.DataFrame) -> pd.DataFrame:
    """
    Create the "Analyse de performance interne" view with fixed columns:
    ID, NOM, DEVISE, TYPE_PART, PORTEFEUILLE, BENCHMARK,
    PTF_BONE, BENCH_ONE, ECART-PERF-BONE, PTF_IFM, BENCH_IFM, ECART-PERF-IFM
    """
    out = df.copy()
    # Map/compute requested columns
    out["PTF_BONE"] = out.get("PF_BONE")
    out["BENCH_ONE"] = out.get("BM_BONE")
    out["ECART-PERF-BONE"] = (out["PTF_BONE"] - out["BENCH_ONE"]) if "PTF_BONE" in out and "BENCH_ONE" in out else np.nan
    out["PTF_IFM"] = out.get("PF_IFM")
    out["BENCH_IFM"] = out.get("BM_IFM")
    out["ECART-PERF-IFM"] = (out["PTF_IFM"] - out["BENCH_IFM"]) if "PTF_IFM" in out and "BENCH_IFM" in out else np.nan

    # Enforce numeric types on perf columns
    for c in ["PTF_BONE", "BENCH_ONE", "ECART-PERF-BONE", "PTF_IFM", "BENCH_IFM", "ECART-PERF-IFM"]:
        if c in out.columns:
            out[c] = pd.to_numeric(out[c], errors="coerce")

    cols = [
        "ID", "NOM", "DEVISE", "TYPE_PART", "PORTEFEUILLE", "BENCHMARK",
        "PTF_BONE", "BENCH_ONE", "ECART-PERF-BONE", "PTF_IFM", "BENCH_IFM", "ECART-PERF-IFM",
    ]
    # Keep only available columns in that order
    cols = [c for c in cols if c in out.columns]
    return out[cols]


def to_attribution_display(df: pd.DataFrame) -> pd.DataFrame:
    """
    Create the "Attribution PF vs BM" view with fixed columns:
    ID, NOM, DEVISE, TYPE_PART, PORTEFEUILLE, BENCHMARK,
    PTF_BONE, PTF_IFM, ECART-PERF-PTF, BENCH_ONE, BENCH_IFM, ECART-PERF-BENCH
    """
    out = df.copy()
    out["PTF_BONE"] = out.get("PF_BONE")
    out["PTF_IFM"] = out.get("PF_IFM")
    out["ECART-PERF-PTF"] = (out["PTF_IFM"] - out["PTF_BONE"]) if "PTF_IFM" in out and "PTF_BONE" in out else np.nan
    out["BENCH_ONE"] = out.get("BM_BONE")
    out["BENCH_IFM"] = out.get("BM_IFM")
    out["ECART-PERF-BENCH"] = (out["BENCH_IFM"] - out["BENCH_ONE"]) if "BENCH_IFM" in out and "BENCH_ONE" in out else np.nan

    for c in ["PTF_BONE", "PTF_IFM", "ECART-PERF-PTF", "BENCH_ONE", "BENCH_IFM", "ECART-PERF-BENCH"]:
        if c in out.columns:
            out[c] = pd.to_numeric(out[c], errors="coerce")

    cols = [
        "ID", "NOM", "DEVISE", "TYPE_PART", "PORTEFEUILLE", "BENCHMARK",
        "PTF_BONE", "PTF_IFM", "ECART-PERF-PTF", "BENCH_ONE", "BENCH_IFM", "ECART-PERF-BENCH",
    ]
    cols = [c for c in cols if c in out.columns]
    return out[cols]

# Patch marker: utils hardening (parsing %, MultiIndex export, ID casts) — OK
TEST_LINE = "utils.py chargé correctement ✅"